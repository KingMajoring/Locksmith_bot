"""Build the Excel sheet sent to a locksmith for their weekly stock check.

Deliberately omits the expected quantity — locksmiths count blind so the
count reflects what's actually there, not what they think we expect.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font

from ..models import WeeklyStockCheck


def build_workbook(weekly_check: WeeklyStockCheck) -> BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock Check"

    ws.append([f"Weekly stock check — {weekly_check.locksmith.name}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Week commencing {weekly_check.week_starting:%d %B %Y}"])
    ws.append([])

    header_row = ws.max_row + 1
    ws.append(["Part code", "Description", "Counted quantity"])
    for cell in ws[header_row]:
        cell.font = Font(bold=True)

    for item in weekly_check.items.order_by("part_code"):
        ws.append([item.part_code, item.part_name, None])

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 36
    ws.column_dimensions["C"].width = 18

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
